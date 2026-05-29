# -*- coding: utf-8 -*-
"""Export 测试用例_硬盘扫描专项.md G: 物理盘表格到 xlsx（对外 / 后端自测）。"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "测试用例_硬盘扫描专项.md"
OUT_PATH = ROOT / "测试用例_硬盘扫描专项.xlsx"
OUT_FALLBACK = ROOT / "测试用例_硬盘扫描专项_物理盘.xlsx"
OUT_BACKEND = ROOT / "测试用例_硬盘扫描专项_后端包.xlsx"

# G: 核心 5 种情况 × 4 种状态（后端自测）
CORE_SITUATIONS: list[tuple[str, str, str]] = [
    ("1", "普通删除 → 回收站保留", "Delete 删除，不清空回收站", "DISK_RECYCLE_001"),
    ("2", "普通删除 → 清空回收站", "Delete 删除后立即清空回收站", "DISK_RECYCLE_002"),
    ("3", "隐藏文件展示/识别（未删除）", "设 Hidden，不删除", "DISK_HIDDEN_003"),
    ("4", "隐藏文件 → 回收站保留", "Hidden + Delete，不清空回收站", "DISK_HIDDEN_004"),
    ("5", "隐藏文件 → 清空回收站", "Hidden + Delete + 清空回收站", "DISK_HIDDEN_005"),
]

OPTIONAL_SITUATIONS: list[tuple[str, str, str]] = [
    ("—", "普通文件永久删除（可选）", "Shift+Delete，不进回收站", "DISK_PERM_003"),
    ("—", "隐藏文件永久删除（可选）", "Hidden + Shift+Delete", "DISK_HIDDEN_006"),
]

# 情况 6：嵌套文件夹路径抽测（Delete→回收站保留，每层 1 条）
FOLDER_DEPTHS: list[tuple[str, str, str, str, str]] = [
    ("1", "1 层", "G:\\nest_L1\\", "DISK_FOLDER_L01", "G:\\nest_L1\\G_nest_L01_<yyMMdd_HHmm>.pptx"),
    ("3", "3 层", "G:\\nest_L1\\nest_L2\\nest_L3\\", "DISK_FOLDER_L03", "G:\\nest_L1\\nest_L2\\nest_L3\\G_nest_L03_<yyMMdd_HHmm>.pptx"),
    ("5", "5 层", "G:\\nest_L1\\…\\nest_L5\\", "DISK_FOLDER_L05", "完整 5 层路径、无截断/乱码"),
    ("10", "10 层", "G:\\nest_L1\\…\\nest_L10\\", "DISK_FOLDER_L10", "完整 10 层路径、无截断/乱码"),
]

STATES: list[tuple[str, str, str]] = [
    ("1", "检出", "扫描结果中能否检索到目标文件"),
    ("2", "原路径", "详情路径应为 G:\\文件名，无乱码"),
    ("3", "预览", "结果页预览正常，无崩溃"),
    ("4", "完整性", "导出/恢复后本地打开，内容完整"),
]

STATE_SUFFIX = ("检出", "原路径", "预览", "完整性")


def parse_table_section(text: str, start_marker: str, end_marker: str) -> tuple[list[str], list[list[str]]]:
    rows: list[list[str]] = []
    in_table = False
    for line in text.splitlines():
        if line.strip().startswith(start_marker):
            in_table = True
            continue
        if in_table and line.strip().startswith(end_marker):
            break
        if not in_table or not line.strip().startswith("|"):
            continue
        if re.match(r"^\|\s*-+", line):
            continue
        cells = [c.strip().replace("`", "") for c in line.strip().strip("|").split("|")]
        rows.append(cells)
    if not rows:
        raise SystemExit(f"No table rows found for section: {start_marker}")
    return rows[0], rows[1:]


def write_sheet(
    ws: Worksheet,
    title: str,
    meta: list[str],
    headers: list[str],
    data: list[list[str]],
) -> None:
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    start = 3
    for i, item in enumerate(meta, start=start):
        ws.cell(row=i, column=1, value=item)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))

    header_row = start + len(meta) + 1
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(data, header_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [20, 8, 44, 50, 30, 18, 12, 20]
    for idx, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.row_dimensions[header_row].height = 28


def _apply_border(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def write_model_sheet(ws: Worksheet) -> None:
    """Sheet: 5 种情况 × 4 种状态 总览（G 盘后端自测）。"""
    ws.title = "5情况x4状态"
    header_fill = PatternFill("solid", fgColor="4472C4")
    sub_fill = PatternFill("solid", fgColor="D9E2F3")
    opt_fill = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = "G: 物理盘 — 测试模型总览（后端自测）"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    meta = [
        "测试盘：G: Test5GB | 情况1～5：G:\\ 根目录 | 情况6：嵌套文件夹路径抽测（1/3/5/10 层，各 1 条）",
        "每种情况含 4 条子用例：检出 / 原路径 / 预览 / 完整性；完整步骤见 sheet「G盘-明细」",
    ]
    for i, line in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)

    # --- 4 种状态说明 ---
    state_title_row = 6
    ws.cell(row=state_title_row, column=1, value="4 种验收状态").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=state_title_row, start_column=1, end_row=state_title_row, end_column=9)

    state_hdr = state_title_row + 1
    for col, h in enumerate(["序号", "状态", "验收要点", "用例后缀"], 1):
        cell = ws.cell(row=state_hdr, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    suffixes = ("-1", "-2", "-3", "-4")
    for ri, (no, name, desc) in enumerate(STATES, start=state_hdr + 1):
        ws.cell(row=ri, column=1, value=no)
        ws.cell(row=ri, column=2, value=name)
        ws.cell(row=ri, column=3, value=desc)
        ws.cell(row=ri, column=4, value=suffixes[int(no) - 1])
        for c in range(1, 5):
            ws.cell(row=ri, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    _apply_border(ws, state_hdr, state_hdr + len(STATES), 1, 4)

    # --- 5 种情况 × 4 状态 矩阵 ---
    matrix_title_row = state_hdr + len(STATES) + 2
    ws.cell(row=matrix_title_row, column=1, value="5 种情况 × 4 种状态 — 用例 ID 矩阵（核心 20 条）").font = Font(
        bold=True, size=12
    )
    ws.merge_cells(start_row=matrix_title_row, start_column=1, end_row=matrix_title_row, end_column=9)

    matrix_hdr = matrix_title_row + 1
    matrix_headers = ["序号", "情况", "操作要点", "检出(-1)", "原路径(-2)", "预览(-3)", "完整性(-4)", "备注"]
    for col, h in enumerate(matrix_headers, 1):
        cell = ws.cell(row=matrix_hdr, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = matrix_hdr + 1
    for no, scene, ops, case_prefix in CORE_SITUATIONS:
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=scene)
        ws.cell(row=row, column=3, value=ops)
        for si, suffix in enumerate(STATE_SUFFIX, start=4):
            ws.cell(row=row, column=si, value=f"{case_prefix}-{si - 3}")
        ws.cell(row=row, column=8, value="P0 核心")
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c <= 3:
                cell.fill = sub_fill
        row += 1
    _apply_border(ws, matrix_hdr, row - 1, 1, 8)

    # --- 情况 6：嵌套文件夹 ---
    folder_fill = PatternFill("solid", fgColor="E2EFDA")
    folder_title_row = row + 1
    ws.cell(row=folder_title_row, column=1, value="情况 6：嵌套文件夹路径抽测（Delete→回收站保留 · 4 条）").font = Font(
        bold=True, size=12
    )
    ws.merge_cells(start_row=folder_title_row, start_column=1, end_row=folder_title_row, end_column=9)

    folder_hdr = folder_title_row + 1
    folder_headers = ["层级", "深度", "目录示例", "用例ID", "预期路径/要点", "备注"]
    for col, h in enumerate(folder_headers, 1):
        cell = ws.cell(row=folder_hdr, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = folder_hdr + 1
    for depth, label, path_example, case_id, expected in FOLDER_DEPTHS:
        ws.cell(row=row, column=1, value=depth)
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=path_example)
        ws.cell(row=row, column=4, value=case_id)
        ws.cell(row=row, column=5, value=expected)
        ws.cell(row=row, column=6, value="P1 抽测")
        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c <= 3:
                cell.fill = folder_fill
        row += 1
    _apply_border(ws, folder_hdr, row - 1, 1, 6)

    # --- 可选扩展 ---
    opt_title_row = row + 1
    ws.cell(row=opt_title_row, column=1, value="可选扩展（永久删除 · 根目录）").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=opt_title_row, start_column=1, end_row=opt_title_row, end_column=9)

    opt_hdr = opt_title_row + 1
    for col, h in enumerate(matrix_headers, 1):
        cell = ws.cell(row=opt_hdr, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = opt_hdr + 1
    for no, scene, ops, case_prefix in OPTIONAL_SITUATIONS:
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=scene)
        ws.cell(row=row, column=3, value=ops)
        for si, suffix in enumerate(STATE_SUFFIX, start=4):
            ws.cell(row=row, column=si, value=f"{case_prefix}-{si - 3}")
        ws.cell(row=row, column=8, value="可选")
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c <= 3:
                cell.fill = opt_fill
        row += 1
    _apply_border(ws, opt_hdr, row - 1, 1, 8)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 32
    for col in "DEFG":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["H"].width = 10
    ws.freeze_panes = ws.cell(row=matrix_hdr + 1, column=1).coordinate


def write_resources_sheet(ws: Worksheet) -> None:
    ws.title = "G盘-资源准备"
    header_fill = PatternFill("solid", fgColor="4472C4")
    ws["A1"] = "G: 测试资源准备（后端）"
    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=14)

    meta = [
        "测试盘：G: Test5GB | 仅操作 G: | 详细说明见：后端_G盘测试资源准备.md",
        "根目录样本：.\\scripts\\stage_user_preview_resources.ps1",
        "嵌套文件夹：.\\scripts\\stage_g_folder_samples.ps1",
    ]
    for i, line in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    hdr = 7
    headers = ["步骤", "脚本/目录", "说明"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = [
        ("1. 清理", "scripts\\clear_gh_delete.ps1", "清空 G:/H: 根目录（可选；stage 脚本也会清 G:）"),
        ("2. 根目录样本", "scripts\\stage_user_preview_resources.ps1", "情况1～5；命名 G_<ext>_<seq>_<yyMMdd_HHmm>.<ext>"),
        ("3. 文件夹样本", "scripts\\stage_g_folder_samples.ps1", "情况6；生成 nest_L1～L10 四层路径样本，不清空整盘"),
        ("4. 模板目录", "scripts\\user_resources\\", "仅复制真实模板；pptx/xlsx/jpg 已有；docx/pdf 建议补充"),
        ("5. 隐藏文件", "PowerShell", '(Get-Item -LiteralPath "G:\\路径" -Force).Attributes = "Hidden"'),
        ("6. 回收站", "系统", "删除类用例前/后注意清空或保留，与用例步骤一致"),
    ]
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, row in enumerate(rows, start=hdr + 1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    tpl_hdr = hdr + len(rows) + 2
    ws.cell(row=tpl_hdr, column=1, value="user_resources 已有模板").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=tpl_hdr, start_column=1, end_row=tpl_hdr, end_column=4)
    tpl_row = tpl_hdr + 1
    tpl_headers = ["类型", "路径", "数量/说明"]
    for col, h in enumerate(tpl_headers, 1):
        cell = ws.cell(row=tpl_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    tpl_data = [
        ("pptx", "scripts\\user_resources\\ppt\\", "2 个"),
        ("xlsx", "scripts\\user_resources\\xlsx\\", "3 个"),
        ("jpg", "scripts\\user_resources\\图片\\", "多张"),
        ("doc/docx/xls/ppt", "—", "需自行放入 user_resources，否则跳过"),
    ]
    for ri, row in enumerate(tpl_data, start=tpl_row + 1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    _apply_border(ws, tpl_row, tpl_row + len(tpl_data), 1, 3)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 48
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1).coordinate


def main() -> None:
    text = MD_PATH.read_text(encoding="utf-8")
    g_headers, g_data = parse_table_section(text, "## 3.", "## 4.")

    wb = Workbook()
    ws_model = wb.active
    write_model_sheet(ws_model)

    ws_g = wb.create_sheet("G盘-明细")
    write_sheet(
        ws_g,
        "G: 物理盘（HDD 5GB）功能用例 — 后端自测",
        [
            "用途：后端自测 | 范围：仅 G: 物理盘（5GB HDD，最稳定）",
            "模型：情况1～5（根目录）+ 情况6（嵌套文件夹）— 见「5情况x4状态」",
            "资源准备：见 sheet「G盘-资源准备」或 后端_G盘测试资源准备.md",
        ],
        g_headers,
        g_data,
    )

    ws_res = wb.create_sheet("G盘-资源准备")
    write_resources_sheet(ws_res)

    for path in (OUT_PATH, OUT_FALLBACK, OUT_BACKEND):
        try:
            wb.save(path)
            print(f"Exported: {path}")
            print(f"Sheets: 5情况x4状态, G盘-明细, G盘-资源准备 | G rows: {len(g_data)}")
            if path != OUT_PATH:
                print("原 xlsx 被占用，请关闭 Excel 后重新导出或改用上述文件")
            return
        except PermissionError:
            continue
    raise SystemExit(f"无法写入 xlsx: {OUT_PATH}")


if __name__ == "__main__":
    main()
